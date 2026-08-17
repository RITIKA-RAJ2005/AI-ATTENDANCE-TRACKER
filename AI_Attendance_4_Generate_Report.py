"""
STEP 4: ATTENDANCE REPORT & DASHBOARD
----------------------------------------
Reads every daily CSV file inside attendance/ and produces a single
consolidated Excel report (attendance_report.xlsx) with:
  - A "Summary" sheet: Present / Late / Absent counts and attendance %
    for each registered person, plus a bar chart.
  - A "Daily Records" sheet: every raw row from every day, combined
    (with a Date column), sorted chronologically.
  - One sheet per day (e.g. "2026-08-16") with that day's raw CSV rows.

This can be run standalone, or imported and called from
AI_Attendance_3_Mark_Attendance.py so the report regenerates
automatically every time a camera session ends.

Requirements (in addition to opencv-contrib-python and numpy):
    pip install openpyxl

Usage:
    python AI_Attendance_4_Generate_Report.py
"""

import os
import csv
import glob
import json
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

ATTENDANCE_DIR = "attendance"
TRAINER_LABELS = os.path.join("trainer", "labels.json")
OUTPUT_PATH = "attendance_report.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Excel sheet names can't exceed 31 characters or contain: \ / ? * [ ]
INVALID_SHEET_CHARS = r'\/?*[]'


def load_roster():
    """The set of everyone who has ever been registered/trained."""
    if not os.path.exists(TRAINER_LABELS):
        return []
    with open(TRAINER_LABELS, "r") as f:
        raw = json.load(f)
    return sorted(raw.values())


def load_daily_files():
    return sorted(glob.glob(os.path.join(ATTENDANCE_DIR, "attendance_*.csv")))


def safe_sheet_name(date_str):
    name = "".join(c for c in date_str if c not in INVALID_SHEET_CHARS)
    return name[:31]


def build_summary(files, roster):
    summary = {name: {"present": 0, "late": 0, "absent": 0} for name in roster}
    daily_rows = []          # combined, with date column, for "Daily Records"
    per_day_rows = {}        # date_str -> list of [Name, Status, Check-In, Check-Out]

    for path in files:
        date_str = os.path.basename(path).replace("attendance_", "").replace(".csv", "")
        per_day_rows[date_str] = []
        with open(path, "r", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = row.get("Name", "")
                status = row.get("Status", "Present")
                checkin = row.get("Check-In", "")
                checkout = row.get("Check-Out", "")

                if name not in summary:
                    summary[name] = {"present": 0, "late": 0, "absent": 0}

                if status == "Present":
                    summary[name]["present"] += 1
                elif status == "Late":
                    summary[name]["late"] += 1
                else:
                    summary[name]["absent"] += 1

                daily_rows.append([date_str, name, status, checkin, checkout])
                per_day_rows[date_str].append([name, status, checkin, checkout])

    return summary, len(files), daily_rows, per_day_rows


def style_header_row(ws, row_idx=1):
    for cell in ws[row_idx]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def generate_report(silent=False):
    """Builds attendance_report.xlsx. Returns True on success, False if there's nothing to report yet."""
    roster = load_roster()
    if not roster:
        if not silent:
            print("[ERROR] No trained model/roster found. Register and train people first.")
        return False

    files = load_daily_files()
    if not files:
        if not silent:
            print("[WARNING] No attendance records found yet in the 'attendance' folder.")
        return False

    summary, total_days, daily_rows, per_day_rows = build_summary(files, roster)

    wb = Workbook()

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Name", "Present", "Late", "Absent", "Total Days", "Attendance %"])
    style_header_row(ws)

    for name in sorted(summary):
        s = summary[name]
        attended = s["present"] + s["late"]
        pct = round((attended / total_days) * 100, 1) if total_days else 0
        ws.append([name, s["present"], s["late"], s["absent"], total_days, pct])

    for col, width in zip("ABCDEF", [22, 10, 8, 10, 12, 14]):
        ws.column_dimensions[col].width = width

    chart = BarChart()
    chart.title = "Attendance % by Person"
    chart.y_axis.title = "Attendance %"
    chart.x_axis.title = "Name"
    data = Reference(ws, min_col=6, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 20
    chart.height = 10
    ws.add_chart(chart, "H2")

    # ---- Daily Records sheet (combined, chronological) ----
    ws2 = wb.create_sheet("Daily Records")
    ws2.append(["Date", "Name", "Status", "Check-In", "Check-Out"])
    style_header_row(ws2)
    for row in daily_rows:
        ws2.append(row)
    for col, width in zip("ABCDE", [14, 22, 10, 12, 12]):
        ws2.column_dimensions[col].width = width

    # ---- One sheet per day ----
    for date_str in sorted(per_day_rows):
        ws_day = wb.create_sheet(safe_sheet_name(date_str))
        ws_day.append(["Name", "Status", "Check-In", "Check-Out"])
        style_header_row(ws_day)
        for row in per_day_rows[date_str]:
            ws_day.append(row)
        for col, width in zip("ABCD", [22, 10, 12, 12]):
            ws_day.column_dimensions[col].width = width

    wb.save(OUTPUT_PATH)
    if not silent:
        print(f"[SUCCESS] Report generated: {OUTPUT_PATH}")
        print(f"[INFO] Covered {total_days} day(s) of attendance for {len(summary)} people.")
    return True


if __name__ == "__main__":
    generate_report()
